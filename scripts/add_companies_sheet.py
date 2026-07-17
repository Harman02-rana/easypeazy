from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent))
from company_links import COMPANY_LINKS

BASE = Path(r"d:\Setup Files\Desktop\easypeazyy")
WB_PATH = BASE / "JobHunterPro.xlsx"

COLUMNS = ["Company", "Official Careers Page", "Official Internship Page", "Official New Grad / University Page", "Company Website"]
WIDTHS = {"Company": 30, "Official Careers Page": 45, "Official Internship Page": 50, "Official New Grad / University Page": 50, "Company Website": 32}

DARK_BLUE = "1F4E78"
BAND_COLOR = "DCE6F1"
WHITE = "FFFFFF"

wb = openpyxl.load_workbook(WB_PATH)
if "Companies" in wb.sheetnames:
    del wb["Companies"]
ws = wb.create_sheet(title="Companies")

ws.append(COLUMNS)
for cell in ws[1]:
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

link_font = Font(color="0563C1", underline="single")
blank_count = {"careers": 0, "internship": 0, "new_grad": 0}

for company in sorted(COMPANY_LINKS.keys(), key=str.lower):
    data = COMPANY_LINKS[company]
    ws.append([company, data["careers"], data["internship"], data["new_grad"], data["website"]])
    r = ws.max_row
    fill = PatternFill(start_color=BAND_COLOR, end_color=BAND_COLOR, fill_type="solid") if r % 2 == 0 else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    for ci in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=r, column=ci)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for ci, key in [(2, "careers"), (3, "internship"), (4, "new_grad"), (5, "website")]:
        cell = ws.cell(row=r, column=ci)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = link_font
    if not data["careers"]:
        blank_count["careers"] += 1
    if not data["internship"]:
        blank_count["internship"] += 1
    if not data["new_grad"]:
        blank_count["new_grad"] += 1

last_row = ws.max_row
table = Table(displayName="CompaniesTable", ref=f"A1:{get_column_letter(len(COLUMNS))}{last_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
ws.add_table(table)
ws.freeze_panes = "A2"

for idx, col_name in enumerate(COLUMNS, start=1):
    max_len = max([len(col_name)] + [len(str(ws.cell(row=r, column=idx).value or "")) for r in range(2, last_row + 1)])
    ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, WIDTHS[col_name])

wb.save(WB_PATH)
print(f"Companies sheet added with {last_row - 1} companies")
print(f"Blank counts (left blank because not verifiable): {blank_count}")
