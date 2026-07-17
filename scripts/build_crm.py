import re
import html
import datetime
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

BASE = Path(r"d:\Setup Files\Desktop\easypeazyy")
OUT_PATH = BASE / "JobHunterPro.xlsx"

RAW_FILES = [
    ("New Grad", BASE / "raw" / "NEW_GRAD_INTL.md"),
    ("Internships", BASE / "raw" / "INTERN_INTL.md"),
]

TODAY = datetime.date.today()

# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

ROW_RE = re.compile(
    r'^\|\s*<a href="([^"]+)"><strong>(.*?)</strong></a>\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|'
    r'\s*<a href="([^"]+)">.*?</a>\s*\|\s*(.*?)\s*\|$'
)
SECTION_RE = re.compile(r'^###\s+(.*)$')
TAG_RE = re.compile(r'<[^>]+>')
PLUS_SUFFIX_RE = re.compile(r'\s*\+\d+\s*$')
AGE_RE = re.compile(r'^(\d+)\s*d$', re.IGNORECASE)

DASH_CHARS = "‐‑‒–—―−"
DASH_NORMALIZE_RE = re.compile(f"[{DASH_CHARS}]")


def clean(text):
    text = TAG_RE.sub("", text)
    text = html.unescape(text)
    text = DASH_NORMALIZE_RE.sub("-", text)
    return re.sub(r"\s+", " ", text).strip()


def extract_country(location):
    loc = location.strip()
    if not loc:
        return ""
    if "," in loc:
        part = loc.rsplit(",", 1)[1]
    elif " - " in loc:
        part = loc.rsplit(" - ", 1)[1]
    else:
        part = loc
    part = PLUS_SUFFIX_RE.sub("", part)
    return part.strip()


def age_to_date(age_text):
    m = AGE_RE.match(age_text.strip())
    if not m:
        return ""
    days = int(m.group(1))
    return (TODAY - datetime.timedelta(days=days)).isoformat()


def parse_file(path, sheet_name):
    section = sheet_name
    rows = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        m = SECTION_RE.match(line)
        if m:
            section = clean(m.group(1))
            continue
        m = ROW_RE.match(line)
        if m:
            company_url, company, position, location, apply_url, age = m.groups()
            rows.append(
                {
                    "Sheet": sheet_name,
                    "Company": clean(company),
                    "CompanyURL": company_url.strip(),
                    "Position": clean(position),
                    "Location": clean(location),
                    "Country": extract_country(clean(location)),
                    "Apply Link": apply_url.strip(),
                    "Date Posted": age_to_date(age),
                    "Source": section,
                }
            )
    return rows


# ---------------------------------------------------------------------------
# STEP 3 -- normalization
# ---------------------------------------------------------------------------

COMPANY_SUFFIX_RE = re.compile(
    r'[,]?\s*\b(LLC|Inc\.?|Incorporated|Corporation|Corp\.?|Ltd\.?|Limited)\.?\s*$',
    re.IGNORECASE,
)

LOCATION_CITY_MAP = {
    "bangalore": "Bengaluru",
    "bombay": "Mumbai",
    "madras": "Chennai",
    "calcutta": "Kolkata",
}

REMOTE_PATTERNS = [
    (re.compile(r'Remote\s*\(\s*US\s*\)', re.IGNORECASE), "Remote - USA"),
    (re.compile(r'Remote\s*\(\s*U\.?S\.?A\.?\s*\)', re.IGNORECASE), "Remote - USA"),
    (re.compile(r'Remote\s*\(\s*UK\s*\)', re.IGNORECASE), "Remote - UK"),
    (re.compile(r'Remote\s*\(\s*India\s*\)', re.IGNORECASE), "Remote - India"),
    (re.compile(r'Remote\s*\(\s*Canada\s*\)', re.IGNORECASE), "Remote - Canada"),
]


def normalize_company(name):
    original = name
    name = COMPANY_SUFFIX_RE.sub('', name).strip()
    return name if name else original


def normalize_location(loc):
    if not loc:
        return loc
    result = loc
    for pattern, replacement in REMOTE_PATTERNS:
        result = pattern.sub(replacement, result)
    for old, new in LOCATION_CITY_MAP.items():
        result = re.sub(rf'\b{old}\b', new, result, flags=re.IGNORECASE)
    return result


def apply_normalization(rows):
    changed = 0
    for row in rows:
        new_company = normalize_company(row["Company"])
        new_location = normalize_location(row["Location"])
        if new_company != row["Company"] or new_location != row["Location"]:
            changed += 1
        row["Company"] = new_company
        row["Location"] = new_location
        row["Country"] = extract_country(new_location)
    return changed


# ---------------------------------------------------------------------------
# STEP 1 -- eligibility classification
# ---------------------------------------------------------------------------

STRONG_EXCLUDE = [
    (r'\bsenior\b', "senior"),
    (r'\bsr\.?\b', "sr"),
    (r'\bstaff\b', "staff"),
    (r'\bprincipal\b', "principal"),
    (r'\bprin\.?\b', "principal-abbr"),
    (r'\bmanager\b', "manager"),
    (r'\bdirector\b', "director"),
    (r'\bhead of\b', "head-of"),
    (r'\bvp\b', "vp"),
    (r'\bvice president\b', "vp"),
    (r'\blead\b', "lead"),
    (r'\barchitect\b', "architect"),
    (r'\bexperienced\b', "experienced"),
    (r'\b[2-9]\+?\s*years?\b', "years-experience"),
]

SOFT_EXCLUDE = [
    (r'\bmid[- ]level\b', "mid-level"),
    (r'\b(ii|iii|iv)\b', "level-2plus-roman"),
    (r'\bengineer\s*[2-9]\b', "engineer-2plus"),
    (r'\bl[5-9]\b', "l5plus"),
]

INCLUDE_PATTERNS = [
    (r'\bintern(ship)?s?\b', "intern"),
    (r'\bco-?op\b', "co-op"),
    (r'\bnew grad(uate)?s?\b', "new-grad"),
    (r'graduate', "graduate"),
    (r'\bentry[- ]level\b', "entry-level"),
    (r'\bjunior\b', "junior"),
    (r'\bassociate\b', "associate"),
    (r'\btrainee\b', "trainee"),
    (r'\bapprentice\b', "apprentice"),
    (r'\buniversity\b', "university"),
    (r'\bcampus\b', "campus"),
    (r'\bearly[- ]career\b', "early-career"),
    (r'\bfresh(er)?\b', "fresher"),
    (r'\balternance\b', "alternance-fr"),
    (r'\bstagiaire\b', "stagiaire-fr"),
    (r'\bi\b', "level-1-roman"),
    (r'\bl4\b', "l4"),
    (r'\b1\b', "level-1-numeral"),
]


def classify(position):
    p = position.lower()
    strong_hits = [tag for pat, tag in STRONG_EXCLUDE if re.search(pat, p)]
    if strong_hits:
        return "EXCLUDE", strong_hits

    soft_hits = [tag for pat, tag in SOFT_EXCLUDE if re.search(pat, p)]
    include_hits = [tag for pat, tag in INCLUDE_PATTERNS if re.search(pat, p)]

    if soft_hits:
        if include_hits:
            return "REVIEW", soft_hits + include_hits
        return "EXCLUDE", soft_hits

    if include_hits:
        return "KEEP", include_hits

    return "REVIEW", []


# ---------------------------------------------------------------------------
# STEP 2 -- dedupe (same company + role -> keep newest)
# ---------------------------------------------------------------------------

def dedupe_same_role(rows):
    best = {}
    order = []
    for row in rows:
        key = (
            row["Company"].strip().lower(),
            re.sub(r"\s+", " ", row["Position"].strip().lower()),
            row["Location"].strip().lower(),
        )
        if key not in best:
            best[key] = row
            order.append(key)
        else:
            current = best[key]
            if row["Date Posted"] and (not current["Date Posted"] or row["Date Posted"] > current["Date Posted"]):
                best[key] = row
    return [best[k] for k in order]


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

DARK_BLUE = "1F4E78"
BAND_COLOR = "DCE6F1"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FONT = Font(bold=True, color=WHITE)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(vertical="center", wrap_text=True, horizontal="left")
BAND_FILL = PatternFill(start_color=BAND_COLOR, end_color=BAND_COLOR, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
HYPERLINK_FONT = Font(color="0563C1", underline="single")


def write_table_sheet(wb, sheet_name, columns, rows, table_name, width_caps, hyperlink_cols=()):
    ws = wb.create_sheet(title=sheet_name)
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for row in rows:
        ws.append([row.get(c, "") for c in columns])
        r = ws.max_row
        fill = BAND_FILL if r % 2 == 0 else WHITE_FILL
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = fill
            cell.alignment = BODY_ALIGN
            if col_name in hyperlink_cols and cell.value:
                cell.hyperlink = cell.value
                cell.font = HYPERLINK_FONT

    last_row = ws.max_row
    last_col_letter = get_column_letter(len(columns))
    table = Table(displayName=table_name, ref=f"A1:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws.freeze_panes = "A2"

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in rows:
            v = str(row.get(col_name, ""))
            if len(v) > max_len:
                max_len = len(v)
        cap = width_caps.get(col_name, 30)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, cap)

    return ws


# ---------------------------------------------------------------------------
# Sheet builders for the CRM-specific sheets
# ---------------------------------------------------------------------------

JOB_COLUMNS = ["Company", "Position", "Country", "Location", "Apply Link", "Date Posted", "Source"]
JOB_WIDTHS = {
    "Company": 34, "Position": 60, "Country": 18, "Location": 34,
    "Apply Link": 55, "Date Posted": 14, "Source": 14,
}

REVIEW_COLUMNS = JOB_COLUMNS + ["Origin Sheet"]
REVIEW_WIDTHS = dict(JOB_WIDTHS, **{"Origin Sheet": 14})

COMPANIES_COLUMNS = [
    "Company", "Category", "Official Careers Page", "Official Internship Page",
    "Official University / Early Careers Page", "Official Company Website", "Country",
    "Typical Internship Hiring Months", "Typical New Graduate Hiring Months",
    "Role Tags", "Roles Offered", "OA Platform", "Remote",
    "India Hiring", "UAE Hiring", "Global Hiring",
    "Visa Sponsorship", "Priority", "Status", "Notes",
]
COMPANIES_WIDTHS = {
    "Company": 30, "Category": 16, "Official Careers Page": 30,
    "Official Internship Page": 30, "Official University / Early Careers Page": 30,
    "Official Company Website": 34, "Country": 16,
    "Typical Internship Hiring Months": 22, "Typical New Graduate Hiring Months": 22,
    "Role Tags": 34, "Roles Offered": 60, "OA Platform": 16,
    "Remote": 10, "India Hiring": 12, "UAE Hiring": 12, "Global Hiring": 12,
    "Visa Sponsorship": 16, "Priority": 10, "Status": 14, "Notes": 24,
}

TRACKER_COLUMNS = [
    "Company", "Role", "Application Link", "Application Date", "Deadline",
    "Status", "Interview Date", "Offer", "Notes",
]
TRACKER_WIDTHS = {
    "Company": 26, "Role": 40, "Application Link": 45, "Application Date": 16,
    "Deadline": 14, "Status": 14, "Interview Date": 16, "Offer": 12, "Notes": 30,
}

GOV_COLUMNS = ["Organization", "Full Name", "Program", "Location", "Apply Link", "Notes"]
GOV_WIDTHS = {"Organization": 14, "Full Name": 45, "Program": 24, "Location": 18, "Apply Link": 30, "Notes": 24}

GOV_ORGS = [
    ("DRDO", "Defence Research and Development Organisation"),
    ("ISRO", "Indian Space Research Organisation"),
    ("MeitY", "Ministry of Electronics and Information Technology"),
    ("AICTE", "All India Council for Technical Education"),
    ("BARC", "Bhabha Atomic Research Centre"),
    ("BEL", "Bharat Electronics Limited"),
    ("HAL", "Hindustan Aeronautics Limited"),
    ("ECIL", "Electronics Corporation of India Limited"),
    ("CDAC", "Centre for Development of Advanced Computing"),
    ("NIC", "National Informatics Centre"),
]

CALENDAR_COLUMNS = ["Month", "Companies Hiring", "Internship / New Grad", "Notes"]
CALENDAR_WIDTHS = {"Month": 14, "Companies Hiring": 40, "Internship / New Grad": 22, "Notes": 30}
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def build_companies_rows(all_rows):
    by_company = defaultdict(lambda: {
        "categories": set(), "url": "", "roles": [], "remote": False,
    })
    for row in all_rows:
        c = by_company[row["Company"]]
        if row["Source"]:
            c["categories"].add(row["Source"])
        if not c["url"] and row.get("CompanyURL"):
            c["url"] = row["CompanyURL"]
        if row["Position"] not in c["roles"]:
            c["roles"].append(row["Position"])
        if "remote" in row["Location"].lower():
            c["remote"] = True

    rows = []
    for company in sorted(by_company.keys(), key=lambda s: s.lower()):
        data = by_company[company]
        roles_offered = "; ".join(data["roles"])
        if len(roles_offered) > 500:
            roles_offered = roles_offered[:497] + "..."
        rows.append({
            "Company": company,
            "Category": ", ".join(sorted(data["categories"])),
            "Official Careers Page": "",
            "Official Internship Page": "",
            "Official University / Early Careers Page": "",
            "Official Company Website": data["url"],
            "Country": "",
            "Typical Internship Hiring Months": "",
            "Typical New Graduate Hiring Months": "",
            "Role Tags": "",
            "Roles Offered": roles_offered,
            "OA Platform": "",
            "Remote": "Yes" if data["remote"] else "",
            "India Hiring": "",
            "UAE Hiring": "",
            "Global Hiring": "",
            "Visa Sponsorship": "",
            "Priority": "",
            "Status": "",
            "Notes": "",
        })
    return rows


def build_dashboard(wb):
    ws = wb.create_sheet(title="Dashboard")
    ws.merge_cells("A1:B1")
    ws["A1"] = "JobHunterPro — 2027 Placement Dashboard"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    metrics = [
        ("Total Companies", "=COUNTA(Companies!A2:A100000)"),
        ("Total Eligible Jobs", "=COUNTA('New Grad'!A2:A100000)+COUNTA(Internships!A2:A100000)"),
        ("Internships", "=COUNTA(Internships!A2:A100000)"),
        ("New Grad", "=COUNTA('New Grad'!A2:A100000)"),
        ("Applications", "=COUNTA('Application Tracker'!A2:A100000)"),
        ("Offers", '=COUNTIF(\'Application Tracker\'!F2:F100000,"Offer")'),
        ("Interviews", '=COUNTIF(\'Application Tracker\'!F2:F100000,"Interview")'),
        ("Rejected", '=COUNTIF(\'Application Tracker\'!F2:F100000,"Rejected")'),
        ("Pending", '=COUNTIF(\'Application Tracker\'!F2:F100000,"Not Applied")'
                    '+COUNTIF(\'Application Tracker\'!F2:F100000,"Applied")'
                    '+COUNTIF(\'Application Tracker\'!F2:F100000,"OA")'),
    ]

    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 3
    for label, formula in metrics:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, color=DARK_BLUE)
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
        val_cell = ws.cell(row=r, column=2, value=formula)
        val_cell.font = Font(bold=True, size=12)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = border
        fill = BAND_FILL if r % 2 == 0 else WHITE_FILL
        ws.cell(row=r, column=1).fill = fill
        val_cell.fill = fill
        r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"
    return ws


def build_application_tracker(wb):
    ws = write_table_sheet(
        wb, "Application Tracker", TRACKER_COLUMNS, [], "ApplicationTrackerTable", TRACKER_WIDTHS,
        hyperlink_cols=(),
    )
    dv = DataValidation(
        type="list",
        formula1='"Not Applied,Applied,OA,Interview,Rejected,Offer"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please choose a value from the list"
    dv.errorTitle = "Invalid Status"
    status_range = "F2:F1000"
    ws.add_data_validation(dv)
    dv.add(status_range)

    status_colors = {
        "Offer": "C6EFCE",
        "Interview": "FFEB9C",
        "OA": "BDD7EE",
        "Applied": "D9D9D9",
        "Rejected": "FFC7CE",
        "Not Applied": "F2F2F2",
    }
    for status, color in status_colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill),
        )
    return ws


def build_government_sheet(wb):
    rows = [
        {"Organization": abbr, "Full Name": full, "Program": "", "Location": "", "Apply Link": "", "Notes": ""}
        for abbr, full in GOV_ORGS
    ]
    write_table_sheet(wb, "Government Internships", GOV_COLUMNS, rows, "GovernmentTable", GOV_WIDTHS)


def build_calendar_sheet(wb):
    rows = [
        {"Month": m, "Companies Hiring": "", "Internship / New Grad": "", "Notes": ""}
        for m in MONTHS
    ]
    write_table_sheet(wb, "Hiring Calendar", CALENDAR_COLUMNS, rows, "CalendarTable", CALENDAR_WIDTHS)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    report_lines = []

    parsed = {}
    for sheet_name, path in RAW_FILES:
        parsed[sheet_name] = parse_file(path, sheet_name)

    all_rows = parsed["New Grad"] + parsed["Internships"]
    changed = apply_normalization(all_rows)
    report_lines.append(f"Normalization: {changed} rows had company/location text standardized")

    buckets = {"New Grad": {"KEEP": [], "REVIEW": []}, "Internships": {"KEEP": [], "REVIEW": []}}
    exclude_count = {"New Grad": 0, "Internships": 0}

    for sheet_name in ("New Grad", "Internships"):
        for row in parsed[sheet_name]:
            verdict, hits = classify(row["Position"])
            if verdict == "EXCLUDE":
                exclude_count[sheet_name] += 1
            else:
                buckets[sheet_name][verdict].append(row)

    review_rows = []
    for sheet_name in ("New Grad", "Internships"):
        for row in buckets[sheet_name]["REVIEW"]:
            r = dict(row)
            r["Origin Sheet"] = sheet_name
            review_rows.append(r)

    new_grad_final = dedupe_same_role(buckets["New Grad"]["KEEP"])
    intern_final = dedupe_same_role(buckets["Internships"]["KEEP"])
    review_final = dedupe_same_role(review_rows)

    dupes_removed = {
        "New Grad": len(buckets["New Grad"]["KEEP"]) - len(new_grad_final),
        "Internships": len(buckets["Internships"]["KEEP"]) - len(intern_final),
        "Needs Review": len(review_rows) - len(review_final),
    }

    report_lines.append(
        f"New Grad: parsed={len(parsed['New Grad'])} excluded={exclude_count['New Grad']} "
        f"review={len(buckets['New Grad']['REVIEW'])} dupes_removed={dupes_removed['New Grad']} "
        f"final_keep={len(new_grad_final)}"
    )
    report_lines.append(
        f"Internships: parsed={len(parsed['Internships'])} excluded={exclude_count['Internships']} "
        f"review={len(buckets['Internships']['REVIEW'])} dupes_removed={dupes_removed['Internships']} "
        f"final_keep={len(intern_final)}"
    )
    report_lines.append(
        f"Needs Review: combined={len(review_rows)} dupes_removed={dupes_removed['Needs Review']} "
        f"final={len(review_final)}"
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_table_sheet(wb, "New Grad", JOB_COLUMNS, new_grad_final, "NewGradTable", JOB_WIDTHS, hyperlink_cols=("Apply Link",))
    write_table_sheet(wb, "Internships", JOB_COLUMNS, intern_final, "InternshipsTable", JOB_WIDTHS, hyperlink_cols=("Apply Link",))
    write_table_sheet(wb, "Needs Review", REVIEW_COLUMNS, review_final, "NeedsReviewTable", REVIEW_WIDTHS, hyperlink_cols=("Apply Link",))

    companies_rows = build_companies_rows(new_grad_final + intern_final + review_final)
    write_table_sheet(wb, "Companies", COMPANIES_COLUMNS, companies_rows, "CompaniesTable", COMPANIES_WIDTHS,
                       hyperlink_cols=("Official Company Website", "Official Careers Page", "Official Internship Page", "Official University / Early Careers Page"))
    report_lines.append(f"Companies: {len(companies_rows)} unique companies")

    build_application_tracker(wb)
    build_government_sheet(wb)
    build_calendar_sheet(wb)
    build_dashboard(wb)

    wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames) - 1))

    wb.save(OUT_PATH)

    print("\n".join(report_lines))
    print(f"\nSheets: {wb.sheetnames}")
    print(f"Saved to {OUT_PATH}")


if __name__ == "__main__":
    main()
