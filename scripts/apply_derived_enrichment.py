import re
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

import sys
sys.path.insert(0, str(Path(__file__).parent))
from category_map import CATEGORY_MAP

BASE = Path(r"d:\Setup Files\Desktop\easypeazyy")
WB_PATH = BASE / "JobHunterPro.xlsx"

with open(BASE / "scripts" / "company_stats.json", encoding="utf-8") as f:
    stats = json.load(f)

# ---------------------------------------------------------------------------
# Role tags (derived purely from our own scraped position titles)
# ---------------------------------------------------------------------------

ROLE_TAG_RULES = [
    (r'\bfull[- ]?stack\b', "Full Stack"),
    (r'\bback[- ]?end\b', "Backend"),
    (r'\bfront[- ]?end\b', "Frontend"),
    (r'\b(android|ios|mobile)\b', "Mobile"),
    (r'\bembedded\b', "Embedded"),
    (r'\bfirmware\b', "Firmware"),
    (r'\bcloud\b', "Cloud"),
    (r'\bdevops\b', "DevOps"),
    (r'\b(security|cyber)\b', "Security"),
    (r'\b(ai|ml|machine learning|artificial intelligence|genai|agentic)\b', "AI / ML"),
    (r'\bdata\b', "Data Science"),
    (r'\b(qa|quality assurance|\bsdet\b|\btest\b|testing)\b', "QA / Testing"),
    (r'\bnetwork(ing)?\b', "Networking"),
    (r'\bsystems?\b', "Systems"),
]


def role_tags_for(positions):
    tags = []
    for pos in positions:
        p = pos.lower()
        for pattern, tag in ROLE_TAG_RULES:
            if re.search(pattern, p) and tag not in tags:
                tags.append(tag)
    if not tags:
        tags = ["Software Engineering"]
    return ", ".join(tags)


# ---------------------------------------------------------------------------
# India / UAE / Global hiring signals (derived from observed posting countries)
# ---------------------------------------------------------------------------

def hiring_signals(countries):
    countries_lower = {c.lower() for c in countries}
    india = "Yes" if "india" in countries_lower else "Unknown"
    uae = "Yes" if ("united arab emirates" in countries_lower or "uae" in countries_lower) else "Unknown"
    glob = "Yes" if len(countries) >= 3 else "Unknown"
    return india, uae, glob


# ---------------------------------------------------------------------------
# Priority score
# ---------------------------------------------------------------------------

def compute_raw_score(company, s, category):
    intern = s["counts"].get("internship", 0)
    newgrad = s["counts"].get("new_grad", 0)
    total = intern + newgrad
    countries = len(s["countries"])

    score = total * 2 + countries * 1.5
    if "Big Tech" in category:
        score += 10
    if "Semiconductor" in category or "AI / ML" in category:
        score += 5
    if "Indian Product" in category:
        score += 3
    return score


records = []
for company, s in stats.items():
    category = CATEGORY_MAP.get(company, "")
    tags = role_tags_for(s["positions"])
    india, uae, glob = hiring_signals(s["countries"])
    score = compute_raw_score(company, s, category)
    records.append({
        "company": company,
        "category": category,
        "tags": tags,
        "india": india,
        "uae": uae,
        "global": glob,
        "score": score,
        "intern": s["counts"].get("internship", 0),
        "newgrad": s["counts"].get("new_grad", 0),
    })

scores_sorted = sorted(r["score"] for r in records)
n = len(scores_sorted)


def percentile(p):
    idx = min(n - 1, int(n * p))
    return scores_sorted[idx]


p90 = percentile(0.90)
p70 = percentile(0.70)
p40 = percentile(0.40)
p15 = percentile(0.15)

print(f"Score distribution: min={scores_sorted[0]:.1f} p15={p15:.1f} p40={p40:.1f} p70={p70:.1f} p90={p90:.1f} max={scores_sorted[-1]:.1f}")


def stars_for(score):
    if score >= p90:
        return "★★★★★"
    if score >= p70:
        return "★★★★☆"
    if score >= p40:
        return "★★★☆☆"
    if score >= p15:
        return "★★☆☆☆"
    return "★☆☆☆☆"


for r in records:
    r["priority"] = stars_for(r["score"])

by_company = {r["company"]: r for r in records}

with open(BASE / "scripts" / "derived_records.json", "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, indent=2)

# ---------------------------------------------------------------------------
# Write into workbook
# ---------------------------------------------------------------------------

wb = openpyxl.load_workbook(WB_PATH)
ws = wb["Companies"]
headers = [c.value for c in ws[1]]
col = {h: i + 1 for i, h in enumerate(headers)}

filled_category = filled_tags = 0
for r in range(2, ws.max_row + 1):
    company = ws.cell(row=r, column=col["Company"]).value
    rec = by_company.get(company)
    if not rec:
        continue
    if rec["category"]:
        ws.cell(row=r, column=col["Category"], value=rec["category"])
        filled_category += 1
    ws.cell(row=r, column=col["Role Tags"], value=rec["tags"])
    filled_tags += 1
    ws.cell(row=r, column=col["India Hiring"], value=rec["india"])
    ws.cell(row=r, column=col["UAE Hiring"], value=rec["uae"])
    ws.cell(row=r, column=col["Global Hiring"], value=rec["global"])
    ws.cell(row=r, column=col["Priority"], value=rec["priority"])

wb.save(WB_PATH)
print(f"Category filled for {filled_category}/{len(records)} companies")
print(f"Role Tags / India / UAE / Global / Priority filled for {filled_tags} companies")
