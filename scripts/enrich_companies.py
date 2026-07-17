import re
import json
from pathlib import Path
from collections import defaultdict

import openpyxl

BASE = Path(r"d:\Setup Files\Desktop\easypeazyy")
WB_PATH = BASE / "JobHunterPro.xlsx"

wb = openpyxl.load_workbook(WB_PATH)

JOB_SHEETS = {
    "New Grad": "new_grad",
    "Internships": "internship",
    "Needs Review": "review",
}

# company -> aggregated stats
stats = defaultdict(lambda: {
    "positions": [],
    "countries": set(),
    "counts": defaultdict(int),  # new_grad / internship / review counts
    "website": "",
})

for sheet_name, kind in JOB_SHEETS.items():
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        company = ws.cell(row=r, column=col["Company"]).value
        if not company:
            continue
        position = ws.cell(row=r, column=col["Position"]).value or ""
        country = ws.cell(row=r, column=col["Country"]).value or ""
        s = stats[company]
        s["positions"].append(position)
        if country:
            s["countries"].add(country)
        s["counts"][kind] += 1

# pull existing website from Companies sheet
ws = wb["Companies"]
headers = [c.value for c in ws[1]]
col = {h: i + 1 for i, h in enumerate(headers)}
for r in range(2, ws.max_row + 1):
    company = ws.cell(row=r, column=col["Company"]).value
    website = ws.cell(row=r, column=col["Official Company Website"]).value
    if company in stats:
        stats[company]["website"] = website or ""

# Rank companies by total posting volume (proxy for relevance) to decide research tier
ranked = sorted(
    stats.items(),
    key=lambda kv: sum(kv[1]["counts"].values()),
    reverse=True,
)

print(f"Total companies: {len(stats)}")
print("\nTop 60 by posting volume:")
for company, s in ranked[:60]:
    total = sum(s["counts"].values())
    countries = len(s["countries"])
    print(f"  {total:3d} postings | {countries:2d} countries | {company}")

# Save aggregated stats as JSON for reuse in later steps
out = {}
for company, s in stats.items():
    out[company] = {
        "positions": s["positions"],
        "countries": sorted(s["countries"]),
        "counts": dict(s["counts"]),
        "website": s["website"],
    }

with open(BASE / "scripts" / "company_stats.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f"\nSaved aggregated stats for {len(out)} companies to scripts/company_stats.json")
