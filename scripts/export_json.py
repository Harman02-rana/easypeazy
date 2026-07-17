import json
import re
from pathlib import Path

import openpyxl

import sys
sys.path.insert(0, str(Path(__file__).parent))
from company_links import COMPANY_LINKS
from category_map import CATEGORY_MAP

BASE = Path(r"d:\Setup Files\Desktop\easypeazyy")
WB_PATH = BASE / "JobHunterPro.xlsx"
OUT_DIR = BASE / "web" / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def slugify(name):
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


# ---------------------------------------------------------------------------
# jobs.json — every row from New Grad + Internships sheets
# ---------------------------------------------------------------------------

wb = openpyxl.load_workbook(WB_PATH)
jobs = []
job_id = 1
company_role_counts = {}  # company -> {"internship": n, "new_grad": n}

for sheet_name, job_type in [("New Grad", "New Grad"), ("Internships", "Internship")]:
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        company = ws.cell(row=r, column=col["Company"]).value
        if not company:
            continue
        position = ws.cell(row=r, column=col["Position"]).value or ""
        country = ws.cell(row=r, column=col["Country"]).value or ""
        location = ws.cell(row=r, column=col["Location"]).value or ""
        apply_link = ws.cell(row=r, column=col["Apply Link"]).value or ""
        date_posted = ws.cell(row=r, column=col["Date Posted"]).value or ""
        source = ws.cell(row=r, column=col["Source"]).value or ""

        jobs.append({
            "id": job_id,
            "company": company,
            "companySlug": slugify(company),
            "position": position,
            "type": job_type,
            "country": country,
            "location": location,
            "applyLink": apply_link,
            "datePosted": date_posted,
            "source": source,
        })
        job_id += 1

        counts = company_role_counts.setdefault(company, {"internship": 0, "new_grad": 0})
        if job_type == "Internship":
            counts["internship"] += 1
        else:
            counts["new_grad"] += 1

with open(OUT_DIR / "jobs.json", "w", encoding="utf-8") as f:
    json.dump(jobs, f, ensure_ascii=False, indent=2)

print(f"jobs.json: {len(jobs)} jobs written")

# ---------------------------------------------------------------------------
# companies.json — the 50 researched companies, enriched with category +
# role counts derived from jobs.json (own data, not invented)
# ---------------------------------------------------------------------------

companies = []
for name in sorted(COMPANY_LINKS.keys(), key=str.lower):
    links = COMPANY_LINKS[name]
    counts = company_role_counts.get(name, {"internship": 0, "new_grad": 0})
    companies.append({
        "slug": slugify(name),
        "name": name,
        "category": CATEGORY_MAP.get(name, ""),
        "officialCareers": links["careers"],
        "internships": links["internship"],
        "newGrad": links["new_grad"],
        "website": links["website"],
        "hiringSeason": "",
        "openInternshipRoles": counts["internship"],
        "openNewGradRoles": counts["new_grad"],
    })

with open(OUT_DIR / "companies.json", "w", encoding="utf-8") as f:
    json.dump(companies, f, ensure_ascii=False, indent=2)

print(f"companies.json: {len(companies)} companies written")

categories = sorted({c["category"] for c in companies if c["category"]})
print("Categories present:", categories)
