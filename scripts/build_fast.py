import re
import html
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

BASE = Path(r"d:\Setup Files\Desktop\easypeazyy")
OUT_PATH = BASE / "JobHunterPro_new.xlsx"

RAW_FILES = [
    ("New Grad", BASE / "raw" / "NEW_GRAD_INTL.md"),
    ("Internships", BASE / "raw" / "INTERN_INTL.md"),
]

TODAY = datetime.date.today()

ROW_RE = re.compile(
    r'^\|\s*<a href="([^"]+)"><strong>(.*?)</strong></a>\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|'
    r'\s*<a href="([^"]+)">.*?</a>\s*\|\s*(.*?)\s*\|$'
)
SECTION_RE = re.compile(r'^###\s+(.*)$')
TAG_RE = re.compile(r'<[^>]+>')
PLUS_SUFFIX_RE = re.compile(r'\s*\+\d+\s*$')
AGE_RE = re.compile(r'^(\d+)\s*d$', re.IGNORECASE)
DASH_CHARS = "‐‑‒–—―−"
DASH_NORMALIZE_RE = re.compile(f"[{DASH_CHARS}]")


def clean(text):
    text = TAG_RE.sub("", text)
    text = html.unescape(text)
    text = DASH_NORMALIZE_RE.sub("-", text)
    return re.sub(r"\s+", " ", text).strip()


def extract_country(location):
    loc = location.strip()
    if not loc:
        return ""
    if "," in loc:
        part = loc.rsplit(",", 1)[1]
    elif " - " in loc:
        part = loc.rsplit(" - ", 1)[1]
    else:
        part = loc
    return PLUS_SUFFIX_RE.sub("", part).strip()


def age_to_date(age_text):
    m = AGE_RE.match(age_text.strip())
    if not m:
        return ""
    return (TODAY - datetime.timedelta(days=int(m.group(1)))).isoformat()


def parse_file(path, sheet_name):
    section = sheet_name
    rows = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        m = SECTION_RE.match(line)
        if m:
            section = clean(m.group(1))
            continue
        m = ROW_RE.match(line)
        if m:
            _company_url, company, position, location, apply_url, age = m.groups()
            rows.append({
                "Company": clean(company),
                "Position": clean(position),
                "Country": extract_country(clean(location)),
                "Location": clean(location),
                "Apply Link": apply_url.strip(),
                "Date Posted": age_to_date(age),
                "Source": section,
            })
    return rows


# Only exclude CLEARLY senior/experienced-only titles. Everything else
# (including ambiguous ones) is kept so no link is dropped.
EXCLUDE_PATTERNS = [
    r'\bsenior\b', r'\bsr\.?\b', r'\bstaff\b',
    r'\bprincipal\b', r'\bprin\.?\b',
    r'\bmanager\b', r'\bdirector\b', r'\bhead of\b',
    r'\bvp\b', r'\bvice president\b', r'\blead\b', r'\barchitect\b',
    r'\bexperienced\b', r'\b[2-9]\+?\s*years?\b',
]


def is_fresher_eligible(position):
    p = position.lower()
    return not any(re.search(pat, p) for pat in EXCLUDE_PATTERNS)


COLUMNS = ["Company", "Position", "Country", "Location", "Apply Link", "Date Posted", "Source"]
WIDTHS = {"Company": 34, "Position": 60, "Country": 18, "Location": 34, "Apply Link": 55, "Date Posted": 14, "Source": 14}

DARK_BLUE = "1F4E78"
BAND_COLOR = "DCE6F1"
WHITE = "FFFFFF"


def write_sheet(wb, sheet_name, rows, table_name):
    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row[c] for c in COLUMNS])
        r = ws.max_row
        fill = PatternFill(start_color=BAND_COLOR, end_color=BAND_COLOR, fill_type="solid") if r % 2 == 0 else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        for ci in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        link_cell = ws.cell(row=r, column=5)
        if row["Apply Link"]:
            link_cell.hyperlink = row["Apply Link"]
            link_cell.font = Font(color="0563C1", underline="single")

    last_row = ws.max_row
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(COLUMNS))}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"

    for idx, col_name in enumerate(COLUMNS, start=1):
        max_len = max([len(col_name)] + [len(str(row[col_name])) for row in rows] or [0])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, WIDTHS[col_name])

    return len(rows)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_links = 0
    for sheet_name, path in RAW_FILES:
        rows = parse_file(path, sheet_name)
        parsed_count = len(rows)
        parsed_links = sum(1 for r in rows if r["Apply Link"])
        eligible = [r for r in rows if is_fresher_eligible(r["Position"])]
        excluded = parsed_count - len(eligible)
        n = write_sheet(wb, sheet_name, eligible, sheet_name.replace(" ", "") + "Table")
        eligible_links = sum(1 for r in eligible if r["Apply Link"])
        total_links += eligible_links
        print(f"{sheet_name}: parsed={parsed_count} (links={parsed_links}) excluded_senior={excluded} kept={n} (links={eligible_links})")

    wb.save(OUT_PATH)
    print(f"Total apply links in final workbook: {total_links}")
    print(f"Saved to {OUT_PATH}")


if __name__ == "__main__":
    main()
