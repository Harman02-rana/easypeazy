import re
import html
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

BASE = Path(r"d:\Setup Files\Desktop\easypeazyy")

SHEETS = [
    ("New Grad", BASE / "raw" / "NEW_GRAD_INTL.md", "NewGradTable"),
    ("Internships", BASE / "raw" / "INTERN_INTL.md", "InternshipsTable"),
]

OUT_PATH = BASE / "JobHunterPro.xlsx"

ROW_RE = re.compile(
    r'^\|\s*<a href="([^"]+)"><strong>(.*?)</strong></a>\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|'
    r'\s*<a href="([^"]+)">.*?</a>\s*\|\s*(.*?)\s*\|$'
)
SECTION_RE = re.compile(r'^###\s+(.*)$')
PLUS_SUFFIX_RE = re.compile(r'\s*\+\d+\s*$')
TAG_RE = re.compile(r'<[^>]+>')
AGE_RE = re.compile(r'^(\d+)\s*d$', re.IGNORECASE)

COLUMNS = ["Company", "Position", "Country", "Location", "Apply Link", "Date Posted", "Source"]
WIDTH_CAPS = {
    "Company": 38,
    "Position": 60,
    "Country": 18,
    "Location": 34,
    "Apply Link": 55,
    "Date Posted": 14,
    "Source": 12,
}

DARK_BLUE = "1F4E78"
BAND_COLOR = "DCE6F1"
WHITE = "FFFFFF"

TODAY = datetime.date.today()


def clean(text):
    text = TAG_RE.sub("", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def extract_country(location):
    loc = location.strip()
    if not loc:
        return ""
    if "," in loc:
        part = loc.rsplit(",", 1)[1]
    elif " - " in loc:
        part = loc.rsplit(" - ", 1)[1]
    else:
        part = loc
    part = PLUS_SUFFIX_RE.sub("", part)
    return part.strip()


def age_to_date(age_text):
    m = AGE_RE.match(age_text.strip())
    if not m:
        return ""
    days = int(m.group(1))
    return (TODAY - datetime.timedelta(days=days)).isoformat()


def parse_file(path, source_default):
    section = source_default
    rows = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        m = SECTION_RE.match(line)
        if m:
            section = clean(m.group(1))
            continue
        m = ROW_RE.match(line)
        if m:
            _company_url, company, position, location, apply_url, age = m.groups()
            company = clean(company)
            position = clean(position)
            location = clean(location)
            apply_url = apply_url.strip()
            country = extract_country(location)
            date_posted = age_to_date(age)
            rows.append(
                {
                    "Company": company,
                    "Position": position,
                    "Country": country,
                    "Location": location,
                    "Apply Link": apply_url,
                    "Date Posted": date_posted,
                    "Source": section,
                }
            )
    return rows


def dedupe(rows):
    seen = set()
    unique_rows = []
    dropped = 0
    for row in rows:
        key = row["Apply Link"].strip().lower().rstrip("/")
        if not key:
            key = (
                row["Company"].strip().lower(),
                row["Position"].strip().lower(),
                row["Location"].strip().lower(),
            )
        if key in seen:
            dropped += 1
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows, dropped


def build_sheet(wb, sheet_name, rows, table_name):
    ws = wb.create_sheet(title=sheet_name)

    ws.append(COLUMNS)
    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    header_font = Font(bold=True, color=WHITE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    hyperlink_font = Font(color="0563C1", underline="single")
    body_align = Alignment(vertical="center", wrap_text=True, horizontal="left")
    band_fill = PatternFill(start_color=BAND_COLOR, end_color=BAND_COLOR, fill_type="solid")
    white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    for row in rows:
        ws.append(
            [
                row["Company"],
                row["Position"],
                row["Country"],
                row["Location"],
                row["Apply Link"],
                row["Date Posted"],
                row["Source"],
            ]
        )
        r = ws.max_row
        is_band = (r % 2 == 0)
        fill = band_fill if is_band else white_fill
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = fill
            cell.alignment = body_align

        link_cell = ws.cell(row=r, column=5)
        if row["Apply Link"]:
            link_cell.hyperlink = row["Apply Link"]
            link_cell.font = hyperlink_font

    last_row = ws.max_row
    last_col_letter = get_column_letter(len(COLUMNS))

    if last_row > 1:
        table_ref = f"A1:{last_col_letter}{last_row}"
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    ws.freeze_panes = "A2"

    for idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row in rows:
            value = str(row[col_name])
            if len(value) > max_len:
                max_len = len(value)
        width = min(max_len + 2, WIDTH_CAPS[col_name])
        ws.column_dimensions[get_column_letter(idx)].width = width

    return len(rows)


def quality_check(sheet_name, rows):
    issues = []

    for i, row in enumerate(rows, start=2):
        if not row["Apply Link"]:
            issues.append(f"Row {i}: missing Apply Link")

        if all(not str(row[c]).strip() for c in COLUMNS):
            issues.append(f"Row {i}: fully empty row")

        for col in COLUMNS:
            val = str(row[col])
            if "<a " in val or "<strong>" in val or "](" in val or val.strip().startswith("[") and "](" in val:
                issues.append(f"Row {i}: leftover markup in {col!r} -> {val[:60]}")

    keys = [row["Apply Link"].strip().lower().rstrip("/") for row in rows if row["Apply Link"]]
    if len(keys) != len(set(keys)):
        issues.append("Duplicate Apply Links remain after dedupe")

    print(f"[{sheet_name}] Quality check: {'PASS' if not issues else 'FAIL'} ({len(rows)} rows)")
    for issue in issues[:20]:
        print("  -", issue)
    return not issues


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary = {}
    for sheet_name, path, table_name in SHEETS:
        raw_rows = parse_file(path, sheet_name)
        unique_rows, dropped = dedupe(raw_rows)
        ok = quality_check(sheet_name, unique_rows)
        count = build_sheet(wb, sheet_name, unique_rows, table_name)
        summary[sheet_name] = (len(raw_rows), dropped, count, ok)

    wb.save(OUT_PATH)

    print()
    for sheet_name, (raw_count, dropped, count, ok) in summary.items():
        print(f"{sheet_name}: parsed={raw_count} dropped_dupes={dropped} final={count} qa_pass={ok}")
    print(f"Saved to {OUT_PATH}")


if __name__ == "__main__":
    main()
